"""Presentation-layer formatting for exports and UI.

This module is the **presentation boundary** for numeric output. It must not be
used when persisting or computing prices.

Data flow
---------
1. **Storage** — SQLite keeps full-precision ``REAL`` values (see
   ``repository.save_price_observation``).
2. **Computation** — ``jobs.py``, ``pack_normalize.py``, and SQL aggregations
   operate on unrounded floats so rankings, gaps, and trends stay accurate.
3. **Presentation** — this module rounds price/quantity/percentage columns to
   two decimals only when writing CSV/Excel files or rendering Streamlit tables.

Do not round in the database or in analytics builders; apply formatting here
(or via ``write_presentation_csv`` / ``write_presentation_excel``) instead.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

PRESENTATION_DECIMALS = 2
EXCEL_NUMBER_FORMAT = "0.00"

# Backward-compatible alias for callers/tests that reference the old name.
DISPLAY_DECIMALS = PRESENTATION_DECIMALS

_EXACT_COLUMNS = frozenset(
    {
        "price",
        "original_price",
        "discounted_price",
        "discount_pct",
        "discount_percentage",
        "normalized_qty",
        "price_per_base",
        "viva_price_per_base",
        "viva_reference_price_per_base",
        "gap_vs_viva_pct",
        "gap_pct",
        "advantage_pct",
        "spread_pct",
        "change_pct",
        "coverage_pct",
        "avg_discount_pct",
        "normalized_basket_total",
        "viva_normalized_total",
        "cheapest_competitor_total",
        "min_price_per_base",
        "max_price_per_base",
        "best_competitor_price_per_base",
        "prev_ppb",
        "viva_total",
        "normalized_subtotal",
        "category_normalized_total",
        "value",
    }
)

_EXCLUDED_COLUMNS = frozenset(
    {
        "line",
        "line_no",
        "run_id",
        "obs_id",
        "store_id",
        "brand_id",
        "basket_item_id",
        "viva_rank",
        "lines_on_promo",
        "promo_count",
        "extraction_ok",
        "extraction_missing_url",
        "missing_url",
        "missing_chain_name",
        "missing_generic_desc",
    }
)

_SUFFIX_PATTERNS = (
    "_price",
    "_original_price",
    "_discount_pct",
    "_price_per_base",
    "_normalized_qty",
    "_gap_vs_viva_pct",
    "_gap_pct",
    "_total",
)

_SPECIAL_SUFFIXES = (
    " - Item Price",
)


def is_presentation_numeric_column(name: str) -> bool:
    """Return True if a column should be shown/exported with fixed decimal places."""
    key = str(name or "").strip()
    if not key:
        return False
    lowered = key.lower()
    if lowered in _EXCLUDED_COLUMNS:
        return False
    if lowered.endswith("_count") or lowered.endswith("_rank"):
        return False
    if lowered in _EXACT_COLUMNS:
        return True
    for suffix in _SPECIAL_SUFFIXES:
        if key.endswith(suffix):
            return True
    for suffix in _SUFFIX_PATTERNS:
        if lowered.endswith(suffix):
            return True
    return False


def round_for_presentation(df: pd.DataFrame, decimals: int = PRESENTATION_DECIMALS) -> pd.DataFrame:
    """Return a copy with presentation numeric columns rounded; computation data unchanged."""
    if df.empty:
        return df.copy()
    out = df.copy()
    for col in out.columns:
        if not is_presentation_numeric_column(str(col)):
            continue
        if not pd.api.types.is_numeric_dtype(out[col]):
            continue
        out[col] = out[col].round(decimals)
    return out


def round_kpi_export_values(df: pd.DataFrame, decimals: int = PRESENTATION_DECIMALS) -> pd.DataFrame:
    """Round numeric cells in heterogeneous weekly-summary ``value`` column."""
    if df.empty or "value" not in df.columns:
        return round_for_presentation(df, decimals=decimals)
    out = round_for_presentation(df, decimals=decimals)
    numeric = pd.to_numeric(out["value"], errors="coerce")
    mask = numeric.notna()
    if mask.any():
        out.loc[mask, "value"] = numeric.loc[mask].round(decimals)
    return out


def apply_excel_presentation_format(path: Path, df: pd.DataFrame) -> None:
    """Set Excel ``0.00`` cell format on presentation numeric columns."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    worksheet = workbook.active
    if worksheet is None:
        workbook.close()
        return

    headers = {str(cell.value): idx for idx, cell in enumerate(worksheet[1], start=1) if cell.value}
    for col_name, col_idx in headers.items():
        if not is_presentation_numeric_column(col_name):
            continue
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = EXCEL_NUMBER_FORMAT

    workbook.save(path)
    workbook.close()


def write_presentation_csv(df: pd.DataFrame, path: Path, **kwargs: Any) -> Path:
    """Write CSV with presentation rounding applied at the export boundary."""
    path.parent.mkdir(parents=True, exist_ok=True)
    rounded = round_for_presentation(df)
    rounded.to_csv(path, index=False, **kwargs)
    return path


def write_presentation_excel(df: pd.DataFrame, path: Path, **kwargs: Any) -> Path:
    """Write Excel with presentation rounding and ``0.00`` cell formatting."""
    path.parent.mkdir(parents=True, exist_ok=True)
    rounded = round_for_presentation(df)
    rounded.to_excel(path, index=False, engine="openpyxl", **kwargs)
    if not rounded.empty:
        apply_excel_presentation_format(path, rounded)
    return path


def streamlit_presentation_column_config(df: pd.DataFrame) -> dict[str, Any]:
    """Build Streamlit ``NumberColumn`` config for presentation numeric columns."""
    try:
        import streamlit as st
    except ImportError:
        return {}

    config: dict[str, Any] = {}
    for col in df.columns:
        if is_presentation_numeric_column(str(col)) and pd.api.types.is_numeric_dtype(df[col]):
            config[str(col)] = st.column_config.NumberColumn(
                format=f"%.{PRESENTATION_DECIMALS}f"
            )
    return config


# Deprecated aliases — prefer the names above.
is_formattable_column = is_presentation_numeric_column
round_display_numerics = round_for_presentation
round_kpi_value_column = round_kpi_export_values
apply_excel_decimal_format = apply_excel_presentation_format
write_csv = write_presentation_csv
write_excel = write_presentation_excel
streamlit_column_config = streamlit_presentation_column_config

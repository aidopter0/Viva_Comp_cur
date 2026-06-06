"""Log OpenAI Chat Completions usage and estimated cost to Excel."""

from __future__ import annotations

import json
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from .settings import OPENAI_USAGE_XLSX

# USD per 1M tokens (standard tier; https://openai.com/api/pricing/)
_MODEL_PRICING: dict[str, tuple[float, float]] = {
    "gpt-5.4-mini": (0.75, 4.50),
    "gpt-5.5": (5.00, 30.00),
}
_DEFAULT_PRICING = _MODEL_PRICING["gpt-5.4-mini"]

_USAGE_COLUMNS = [
    "timestamp",
    "operation",
    "model",
    "context",
    "prompt_tokens",
    "completion_tokens",
    "total_tokens",
    "input_cost_usd",
    "output_cost_usd",
    "total_cost_usd",
]


def _pricing_for_model(model: str) -> tuple[float, float]:
    key = (model or "").strip().lower()
    for name, rates in _MODEL_PRICING.items():
        if key == name.lower():
            return rates
    return _DEFAULT_PRICING


def _estimate_cost(model: str, prompt_tokens: int, completion_tokens: int) -> tuple[float, float, float]:
    input_rate, output_rate = _pricing_for_model(model)
    input_cost = prompt_tokens / 1_000_000 * input_rate
    output_cost = completion_tokens / 1_000_000 * output_rate
    return input_cost, output_cost, input_cost + output_cost


def _format_context(context: dict[str, Any] | None) -> str:
    if not context:
        return ""
    parts = [f"{k}={v}" for k, v in context.items()]
    if parts:
        return ", ".join(parts)
    return json.dumps(context, ensure_ascii=False)


def log_completion(
    resp: Any,
    *,
    operation: str,
    model: str,
    context: dict[str, Any] | None = None,
) -> None:
    """Append one usage row to aiuse/openai_usage.xlsx. Never raises."""
    try:
        usage = getattr(resp, "usage", None)
        prompt_tokens = int(getattr(usage, "prompt_tokens", 0) or 0)
        completion_tokens = int(getattr(usage, "completion_tokens", 0) or 0)
        total_tokens = int(getattr(usage, "total_tokens", 0) or 0)
        if total_tokens == 0 and (prompt_tokens or completion_tokens):
            total_tokens = prompt_tokens + completion_tokens

        input_cost, output_cost, total_cost = _estimate_cost(
            model, prompt_tokens, completion_tokens
        )
        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            operation,
            model,
            _format_context(context),
            prompt_tokens,
            completion_tokens,
            total_tokens,
            round(input_cost, 6),
            round(output_cost, 6),
            round(total_cost, 6),
        ]

        path = OPENAI_USAGE_XLSX
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.is_file():
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "usage"
            ws.append(_USAGE_COLUMNS)
        ws.append(row)
        wb.save(path)
    except Exception:
        pass

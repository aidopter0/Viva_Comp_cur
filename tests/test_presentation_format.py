from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from viva_tracker.presentation_format import (
    apply_excel_presentation_format,
    is_presentation_numeric_column,
    round_for_presentation,
    write_presentation_excel,
)


def test_is_presentation_numeric_column_matches_price_fields():
    assert is_presentation_numeric_column("price")
    assert is_presentation_numeric_column("original_price")
    assert is_presentation_numeric_column("discount_pct")
    assert is_presentation_numeric_column("normalized_qty")
    assert is_presentation_numeric_column("price_per_base")
    assert is_presentation_numeric_column("Viva_Arjan_price")
    assert is_presentation_numeric_column("Viva_Arjan_price_per_base")
    assert is_presentation_numeric_column("Geant_Arjan - Item Price")


def test_is_presentation_numeric_column_excludes_identifiers_and_counts():
    assert not is_presentation_numeric_column("line")
    assert not is_presentation_numeric_column("line_no")
    assert not is_presentation_numeric_column("run_id")
    assert not is_presentation_numeric_column("viva_rank")
    assert not is_presentation_numeric_column("lines_on_promo")
    assert not is_presentation_numeric_column("promo_count")


def test_round_for_presentation_rounds_price_columns_only():
    df = pd.DataFrame(
        {
            "line_no": [1, 2],
            "price": [3.666666, 10.0],
            "original_price": [4.9999, None],
            "discount_pct": [12.3456, 0.0],
            "normalized_qty": [1.234567, 2.0],
            "price_per_base": [3.3333333, 5.5555555],
            "Viva_Arjan_price_per_base": [7.777777, 8.888888],
            "lines_on_promo": [3, 5],
        }
    )
    out = round_for_presentation(df)

    assert out.loc[0, "line_no"] == 1
    assert out.loc[0, "lines_on_promo"] == 3
    assert out.loc[0, "price"] == 3.67
    assert out.loc[0, "original_price"] == 5.0
    assert out.loc[0, "discount_pct"] == 12.35
    assert out.loc[0, "normalized_qty"] == 1.23
    assert out.loc[0, "price_per_base"] == 3.33
    assert out.loc[0, "Viva_Arjan_price_per_base"] == 7.78
    assert pd.isna(out.loc[1, "original_price"])


def test_write_presentation_excel_applies_two_decimal_format(tmp_path: Path):
    df = pd.DataFrame(
        {
            "line_no": [1],
            "price": [5.0],
            "discount_pct": [12.3],
        }
    )
    path = tmp_path / "prices.xlsx"
    write_presentation_excel(df, path)

    workbook = load_workbook(path)
    worksheet = workbook.active
    price_cell = worksheet.cell(row=2, column=2)
    discount_cell = worksheet.cell(row=2, column=3)
    line_cell = worksheet.cell(row=2, column=1)

    assert price_cell.number_format == "0.00"
    assert discount_cell.number_format == "0.00"
    assert line_cell.number_format != "0.00"
    workbook.close()


def test_apply_excel_presentation_format_skips_non_numeric_cells(tmp_path: Path):
    df = pd.DataFrame({"price": [1.23], "note": ["ok"]})
    path = tmp_path / "mixed.xlsx"
    df.to_excel(path, index=False, engine="openpyxl")
    apply_excel_presentation_format(path, df)

    workbook = load_workbook(path)
    worksheet = workbook.active
    assert worksheet.cell(row=2, column=1).number_format == "0.00"
    assert worksheet.cell(row=2, column=2).number_format != "0.00"
    workbook.close()
